"""Generate templates/ADV-profiles.xlsx, the export template of ADV-Explorer.

The column order must match adv::statsexport::profileTemplateColumns() in
src/core/ProfileStatsExport.cpp: row 1 holds the title, row 2 the column
headers, data rows start at row 3. The last two columns receive formulas
(velocity magnitude and horizontal direction) written by the application.

Run with: mamba run -n adv-explorer python tools/make_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    "profile",
    "x (m)", "y (m)", "z (m)",
    "h (m)", "z/h (-)",
    "u mean (m/s)", "u std (m/s)", "u skewness (-)", "u kurtosis (-)",
    "v mean (m/s)", "v std (m/s)", "v skewness (-)", "v kurtosis (-)",
    "w mean (m/s)", "w std (m/s)", "w skewness (-)", "w kurtosis (-)",
    "u'v' (m^2/s^2)", "u'w' (m^2/s^2)", "v'w' (m^2/s^2)",
    "TKE (m^2/s^2)", "eps (m^2/s^3)",
    "U magnitude (m/s)", "direction (deg)",
]


def main() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ADV-profiles"

    title_font = Font(bold=True, size=14, color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    formula_fill = PatternFill("solid", fgColor="DDEBF7")

    sheet.cell(row=1, column=1,
               value="ADV-Explorer vertical profile statistics").font = title_font
    sheet.merge_cells(start_row=1, start_column=1, end_row=1,
                      end_column=len(COLUMNS))

    for col, name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=2, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        width = 12 if col > 1 else 16
        sheet.column_dimensions[get_column_letter(col)].width = width

    # mark the two evaluated columns (filled with formulas by the application:
    # magnitude = SQRT(u_mean^2 + v_mean^2 + w_mean^2),
    # direction = DEGREES(ATAN2(u_mean, v_mean)) relative to the bulk flow x axis)
    for col in (len(COLUMNS) - 1, len(COLUMNS)):
        sheet.cell(row=2, column=col).fill = formula_fill
        sheet.cell(row=2, column=col).font = Font(bold=True, color="1F4E79")

    sheet.freeze_panes = "A3"

    out = Path(__file__).resolve().parent.parent / "templates" / "ADV-profiles.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out)
    print(f"Template written to {out}")


if __name__ == "__main__":
    main()
